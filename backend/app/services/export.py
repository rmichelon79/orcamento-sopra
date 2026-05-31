"""Exporta a grade orçamentária para XLSX (openpyxl).

Reusa as estruturas calculadas por `calcular_grade` e `calcular_consolidado`.
"""
from __future__ import annotations

from datetime import datetime
from io import BytesIO
from typing import Iterable

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

from app.schemas.orcamento import (
    GradeConsolidadaResponse,
    GradeNode,
    GradeResponse,
)

MESES = ["Jan", "Fev", "Mar", "Abr", "Mai", "Jun",
         "Jul", "Ago", "Set", "Out", "Nov", "Dez"]

FORMATO_BRL = '_-"R$" * #,##0.00_-;[Red]-"R$" * #,##0.00_-;_-"R$" * "-"??_-;_-@_-'

_FILL_SINTETICA = PatternFill(fill_type="solid", fgColor="F3F4F6")
_FILL_TOTAL = PatternFill(fill_type="solid", fgColor="1F2937")
_FILL_HEADER = PatternFill(fill_type="solid", fgColor="E5E7EB")
_FONT_BOLD = Font(bold=True)
_FONT_TOTAL = Font(bold=True, color="FFFFFF")
_FONT_HEADER = Font(bold=True)
_BORDER_THIN = Side(border_style="thin", color="D1D5DB")
_BORDER_ALL = Border(left=_BORDER_THIN, right=_BORDER_THIN,
                    top=_BORDER_THIN, bottom=_BORDER_THIN)


def _flatten(nodes: list[GradeNode], depth: int = 0) -> Iterable[tuple[int, GradeNode]]:
    """Gera (depth, node) pra cada conta na árvore (DFS, ordenada por `ordem`)."""
    for n in sorted(nodes, key=lambda x: x.ordem):
        yield depth, n
        yield from _flatten(n.filhas, depth + 1)


def _escrever_grade(
    ws,
    titulo: str,
    arvore: list[GradeNode],
    totais_mes,
    total_geral,
) -> None:
    """Preenche a planilha `ws` com cabeçalho + grade + linha total."""
    # Linha 1: título
    ws.cell(row=1, column=1, value=titulo).font = Font(bold=True, size=12)
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=15)

    # Linha 2: data de geração
    ws.cell(
        row=2, column=1,
        value=f"Gerado em {datetime.now().strftime('%d/%m/%Y %H:%M')}"
    ).font = Font(italic=True, color="6B7280", size=9)
    ws.merge_cells(start_row=2, start_column=1, end_row=2, end_column=15)

    # Linha 4: cabeçalho de colunas
    HEADER_ROW = 4
    headers = ["Código", "Conta"] + MESES + ["Total"]
    for col, h in enumerate(headers, start=1):
        c = ws.cell(row=HEADER_ROW, column=col, value=h)
        c.font = _FONT_HEADER
        c.fill = _FILL_HEADER
        c.alignment = Alignment(horizontal="center" if col >= 3 else "left")
        c.border = _BORDER_ALL

    # Linhas de contas
    row = HEADER_ROW + 1
    for depth, n in _flatten(arvore):
        is_sint = n.natureza == "sintetica"
        ws.cell(row=row, column=1, value=n.codigo)
        # Indenta o nome com espaços (não-quebráveis)
        nome_indentado = (" " * (depth * 3)) + n.nome
        ws.cell(row=row, column=2, value=nome_indentado)
        for m in range(12):
            v = float(n.valores[m])
            cell = ws.cell(row=row, column=3 + m, value=v)
            cell.number_format = FORMATO_BRL
        cell_total = ws.cell(row=row, column=15, value=float(n.total))
        cell_total.number_format = FORMATO_BRL

        if is_sint:
            for col in range(1, 16):
                cell = ws.cell(row=row, column=col)
                cell.font = _FONT_BOLD
                cell.fill = _FILL_SINTETICA
        row += 1

    # Linha de Total Geral
    ws.cell(row=row, column=2, value="Total geral")
    for m in range(12):
        ws.cell(row=row, column=3 + m, value=float(totais_mes[m])).number_format = FORMATO_BRL
    ws.cell(row=row, column=15, value=float(total_geral)).number_format = FORMATO_BRL
    for col in range(1, 16):
        cell = ws.cell(row=row, column=col)
        cell.font = _FONT_TOTAL
        cell.fill = _FILL_TOTAL

    # Larguras
    ws.column_dimensions["A"].width = 14
    ws.column_dimensions["B"].width = 45
    for i in range(3, 15):
        ws.column_dimensions[get_column_letter(i)].width = 16
    ws.column_dimensions["O"].width = 18

    # Freeze pane: trava colunas A/B e linhas até o cabeçalho
    ws.freeze_panes = "C5"


def gerar_xlsx_individual(
    grade: GradeResponse, empreendimento_codigo: str, empreendimento_nome: str
) -> bytes:
    wb = Workbook()
    ws = wb.active
    ws.title = empreendimento_codigo[:31]  # XLSX limita 31 chars
    titulo = (
        f"Orçamento — {empreendimento_codigo} {empreendimento_nome} · "
        f"{grade.orcamento.ano}/v{grade.orcamento.versao} · {grade.orcamento.status}"
    )
    _escrever_grade(ws, titulo, grade.arvore, grade.totais_mes, grade.total_geral)
    buf = BytesIO()
    wb.save(buf)
    return buf.getvalue()


def gerar_xlsx_consolidado(
    grade: GradeConsolidadaResponse, codigos_incluidos: list[str]
) -> bytes:
    wb = Workbook()
    ws = wb.active
    ws.title = "Consolidado"
    titulo = (
        f"Orçamento Consolidado · {grade.ano} · "
        f"soma de {len(codigos_incluidos)} empreendimentos: {', '.join(codigos_incluidos)}"
    )
    _escrever_grade(ws, titulo, grade.arvore, grade.totais_mes, grade.total_geral)
    buf = BytesIO()
    wb.save(buf)
    return buf.getvalue()


# ────────────────────────────────────────────────────────────────────────────
# Export Markdown (estrutura nested pra XMind / outros mind maps)
# ────────────────────────────────────────────────────────────────────────────

def _ordem_natural(codigo: str) -> tuple[int, ...]:
    return tuple(int(p) for p in codigo.split("."))


def _no_md(node: GradeNode, depth: int, lines: list[str], eh_raiz: bool) -> None:
    indent = "  " * depth
    badge = ""
    if eh_raiz:
        if node.tipo_orcamentario == "entrada":
            badge = " [entrada]"
        else:
            badge = " [saída]"
    lines.append(f"{indent}- {node.codigo} {node.nome}{badge}")
    for filha in sorted(node.filhas, key=lambda n: _ordem_natural(n.codigo)):
        _no_md(filha, depth + 1, lines, eh_raiz=False)


def gerar_md_individual(
    grade: GradeResponse, empreendimento_codigo: str, empreendimento_nome: str
) -> str:
    lines = [
        f"# Orçamento — {empreendimento_codigo} {empreendimento_nome} · "
        f"{grade.orcamento.ano}/v{grade.orcamento.versao} · {grade.orcamento.status}",
        "",
    ]
    for raiz in sorted(grade.arvore, key=lambda n: _ordem_natural(n.codigo)):
        _no_md(raiz, 0, lines, eh_raiz=True)
    return "\n".join(lines) + "\n"


def gerar_md_consolidado(
    grade: GradeConsolidadaResponse, codigos_incluidos: list[str]
) -> str:
    lines = [
        f"# Orçamento Consolidado · {grade.ano} · "
        f"soma de {len(codigos_incluidos)} empreendimentos: {', '.join(codigos_incluidos)}",
        "",
    ]
    for raiz in sorted(grade.arvore, key=lambda n: _ordem_natural(n.codigo)):
        _no_md(raiz, 0, lines, eh_raiz=True)
    return "\n".join(lines) + "\n"
